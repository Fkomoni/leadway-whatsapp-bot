"""
Leadway Health — Supervisor Dashboard
Run:  python dashboard.py
Then: http://localhost:5050/

Environment variables:
  SUPERVISOR_PASSWORD   plain-text password for supervisor login  (default: leadway2026)
  DASHBOARD_SECRET      Flask session secret                      (default: change-me)
  DASHBOARD_PORT        port to listen on                         (default: 5050)
"""
import csv
import io
import os
from datetime import datetime, timedelta

from flask import (
    Flask, Response, redirect, render_template_string,
    request, session, url_for,
)

from analytics import get_stats, _connect, categorise_tools

app = Flask(__name__)
app.secret_key = os.getenv("DASHBOARD_SECRET", "leadway-dashboard-secret")

SUPERVISOR_PASSWORD = os.getenv("SUPERVISOR_PASSWORD", "leadway2026")

# ─── Auth helpers ─────────────────────────────────────────────────────────────

def logged_in() -> bool:
    return session.get("supervisor") is True


def _week_bounds(offset: int = 0):
    """Return (start, end) ISO strings for Mon–Sun of week offset from today."""
    today = datetime.utcnow().date()
    monday = today - timedelta(days=today.weekday()) + timedelta(weeks=offset)
    sunday = monday + timedelta(days=6)
    return monday.isoformat() + "T00:00:00", sunday.isoformat() + "T23:59:59"


# ─── Templates ────────────────────────────────────────────────────────────────

_LOGIN_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Leadway Health — Supervisor Login</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body { background: #f0f4f8; }
    .card { border: none; border-radius: 12px; box-shadow: 0 4px 24px rgba(0,0,0,.08); }
    .btn-primary { background: #00796b; border-color: #00796b; }
    .btn-primary:hover { background: #005f56; border-color: #005f56; }
    .brand { color: #00796b; font-weight: 700; font-size: 1.3rem; }
  </style>
</head>
<body class="d-flex align-items-center justify-content-center min-vh-100">
  <div class="card p-4" style="width:360px">
    <div class="text-center mb-4">
      <div class="brand">Leadway Health</div>
      <div class="text-muted small">Supervisor Portal</div>
    </div>
    {% if error %}<div class="alert alert-danger py-2">{{ error }}</div>{% endif %}
    <form method="post">
      <div class="mb-3">
        <label class="form-label">Password</label>
        <input type="password" name="password" class="form-control" autofocus required>
      </div>
      <button class="btn btn-primary w-100">Sign in</button>
    </form>
  </div>
</body>
</html>
"""

_DASHBOARD_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Leadway Health — Weekly Report</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
  <style>
    :root { --brand: #00796b; --brand-light: #e0f2f1; }
    body  { background: #f5f7fa; font-size: .9rem; }
    .navbar { background: var(--brand) !important; }
    .navbar-brand { color: #fff !important; font-weight: 700; }
    .kpi-card { border: none; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,.07); }
    .kpi-card .icon { width: 44px; height: 44px; border-radius: 10px;
                      display:flex;align-items:center;justify-content:center;font-size:1.3rem; }
    .kpi-value { font-size: 1.8rem; font-weight: 700; line-height:1; }
    .chart-card { border: none; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,.07);
                  background: #fff; }
    .table thead th { background: var(--brand-light); font-weight: 600; }
    .badge-pos { background: #d4edda; color: #155724; }
    .badge-neu { background: #fff3cd; color: #856404; }
    .badge-neg { background: #f8d7da; color: #721c24; }
    .btn-brand { background: var(--brand); border-color: var(--brand); color:#fff; }
    .btn-brand:hover { background:#005f56; border-color:#005f56; color:#fff; }
    .week-nav .btn { border-radius: 8px; }
  </style>
</head>
<body>

<nav class="navbar navbar-expand-lg mb-4">
  <div class="container-fluid px-4">
    <span class="navbar-brand"><i class="bi bi-heart-pulse me-2"></i>Leadway Health — Supervisor Report</span>
    <div class="ms-auto d-flex gap-2 align-items-center">
      <a href="{{ url_for('download_csv', start=start, end=end) }}"
         class="btn btn-sm btn-outline-light"><i class="bi bi-filetype-csv me-1"></i>CSV</a>
      <a href="{{ url_for('download_excel', start=start, end=end) }}"
         class="btn btn-sm btn-outline-light"><i class="bi bi-file-earmark-excel me-1"></i>Excel</a>
      <a href="{{ url_for('logout') }}" class="btn btn-sm btn-outline-light">
        <i class="bi bi-box-arrow-right me-1"></i>Logout</a>
    </div>
  </div>
</nav>

<div class="container-fluid px-4">

  <!-- Week navigation -->
  <div class="d-flex align-items-center gap-3 mb-4 week-nav">
    <a href="?week={{ week_offset - 1 }}" class="btn btn-sm btn-outline-secondary">
      <i class="bi bi-chevron-left"></i> Prev week
    </a>
    <span class="fw-semibold">
      <i class="bi bi-calendar3 me-1 text-muted"></i>
      {{ start[:10] }} — {{ end[:10] }}
    </span>
    {% if week_offset < 0 %}
    <a href="?week={{ week_offset + 1 }}" class="btn btn-sm btn-outline-secondary">
      Next week <i class="bi bi-chevron-right"></i>
    </a>
    {% else %}
    <span class="btn btn-sm btn-outline-secondary disabled">Next week <i class="bi bi-chevron-right"></i></span>
    {% endif %}
    {% if week_offset != 0 %}
    <a href="?" class="btn btn-sm btn-brand">This week</a>
    {% endif %}
  </div>

  <!-- KPI row -->
  <div class="row g-3 mb-4">
    <div class="col-6 col-md-3">
      <div class="card kpi-card p-3 h-100">
        <div class="d-flex justify-content-between align-items-start">
          <div>
            <div class="text-muted small mb-1">Total Conversations</div>
            <div class="kpi-value">{{ kpi.total_sessions or 0 }}</div>
          </div>
          <div class="icon" style="background:#e0f2f1;color:var(--brand)"><i class="bi bi-chat-dots"></i></div>
        </div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card kpi-card p-3 h-100">
        <div class="d-flex justify-content-between align-items-start">
          <div>
            <div class="text-muted small mb-1">Total Requests</div>
            <div class="kpi-value">{{ kpi.total_requests or 0 }}</div>
          </div>
          <div class="icon" style="background:#e3f2fd;color:#1565c0"><i class="bi bi-send"></i></div>
        </div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card kpi-card p-3 h-100">
        <div class="d-flex justify-content-between align-items-start">
          <div>
            <div class="text-muted small mb-1">Avg Response Time</div>
            <div class="kpi-value">{{ kpi.avg_rt_s or 0 }}<span class="fs-6 text-muted">s</span></div>
          </div>
          <div class="icon" style="background:#fff3e0;color:#e65100"><i class="bi bi-stopwatch"></i></div>
        </div>
      </div>
    </div>
    <div class="col-6 col-md-3">
      <div class="card kpi-card p-3 h-100">
        <div class="d-flex justify-content-between align-items-start">
          <div>
            <div class="text-muted small mb-1">Customer Sentiment</div>
            <div class="kpi-value {% if (kpi.avg_sentiment or 0) > 0.2 %}text-success
                                    {% elif (kpi.avg_sentiment or 0) < -0.2 %}text-danger
                                    {% else %}text-warning{% endif %}">
              {% if (kpi.avg_sentiment or 0) > 0.2 %}😊
              {% elif (kpi.avg_sentiment or 0) < -0.2 %}😞
              {% else %}😐{% endif %}
              <span class="fs-6">{{ "%.2f"|format(kpi.avg_sentiment or 0) }}</span>
            </div>
          </div>
          <div class="icon" style="background:#fce4ec;color:#c62828"><i class="bi bi-emoji-smile"></i></div>
        </div>
      </div>
    </div>
  </div>

  <!-- Charts row 1 -->
  <div class="row g-3 mb-4">
    <div class="col-md-7">
      <div class="chart-card p-3 h-100">
        <div class="fw-semibold mb-3">Daily Request Volume</div>
        <canvas id="lineChart" height="90"></canvas>
      </div>
    </div>
    <div class="col-md-5">
      <div class="chart-card p-3 h-100">
        <div class="fw-semibold mb-3">Requests by Agent</div>
        <canvas id="agentBar" height="130"></canvas>
      </div>
    </div>
  </div>

  <!-- Charts row 2 -->
  <div class="row g-3 mb-4">
    <div class="col-md-4">
      <div class="chart-card p-3 h-100">
        <div class="fw-semibold mb-3">Customer Sentiment</div>
        <canvas id="sentimentDoughnut" height="160"></canvas>
        <div class="d-flex justify-content-center gap-3 mt-2 small">
          <span><span class="badge badge-pos px-2 py-1 me-1">😊</span>Positive: {{ kpi.pos_count or 0 }}</span>
          <span><span class="badge badge-neu px-2 py-1 me-1">😐</span>Neutral: {{ kpi.neu_count or 0 }}</span>
          <span><span class="badge badge-neg px-2 py-1 me-1">😞</span>Negative: {{ kpi.neg_count or 0 }}</span>
        </div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="chart-card p-3 h-100">
        <div class="fw-semibold mb-3">Request Type Breakdown</div>
        <canvas id="typeBar" height="160"></canvas>
      </div>
    </div>
    <div class="col-md-4">
      <div class="chart-card p-3 h-100">
        <div class="fw-semibold mb-3">Peak Hours (UTC)</div>
        <canvas id="hourBar" height="160"></canvas>
      </div>
    </div>
  </div>

  <!-- Agent performance table -->
  <div class="card chart-card mb-4">
    <div class="card-body">
      <div class="fw-semibold mb-3">Agent Performance</div>
      <div class="table-responsive">
        <table class="table table-sm table-hover mb-0 align-middle">
          <thead>
            <tr>
              <th>Agent</th>
              <th class="text-center">Conversations</th>
              <th class="text-center">Requests</th>
              <th class="text-center">Avg Response Time</th>
              <th class="text-center">Avg Sentiment</th>
              <th class="text-center">Escalations</th>
            </tr>
          </thead>
          <tbody>
            {% for a in agents %}
            <tr>
              <td><i class="bi bi-robot me-1 text-muted"></i>{{ a.agent_name }}</td>
              <td class="text-center">{{ a.sessions }}</td>
              <td class="text-center fw-semibold">{{ a.requests }}</td>
              <td class="text-center">{{ a.avg_rt_s }}s</td>
              <td class="text-center">
                {% if (a.avg_sentiment or 0) > 0.2 %}<span class="badge badge-pos">😊 Positive</span>
                {% elif (a.avg_sentiment or 0) < -0.2 %}<span class="badge badge-neg">😞 Negative</span>
                {% else %}<span class="badge badge-neu">😐 Neutral</span>{% endif %}
              </td>
              <td class="text-center">
                {% if a.escalations %}<span class="text-danger fw-semibold">{{ a.escalations }}</span>
                {% else %}<span class="text-muted">0</span>{% endif %}
              </td>
            </tr>
            {% else %}
            <tr><td colspan="6" class="text-center text-muted py-3">No data for this period</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Request log -->
  <div class="card chart-card mb-5">
    <div class="card-body">
      <div class="fw-semibold mb-3">Request Log <span class="text-muted fw-normal small">(last 200)</span></div>
      <div class="table-responsive">
        <table class="table table-sm table-hover mb-0 align-middle">
          <thead>
            <tr>
              <th>Time (UTC)</th>
              <th>Agent</th>
              <th>Member</th>
              <th>Request</th>
              <th class="text-center">Response Time</th>
              <th class="text-center">Sentiment</th>
              <th>Service</th>
            </tr>
          </thead>
          <tbody>
            {% for r in log %}
            <tr>
              <td class="text-nowrap text-muted small">{{ r.requested_at[:16].replace("T"," ") }}</td>
              <td>{{ r.agent_name }}</td>
              <td class="text-muted small">{{ r.member_name or r.member_phone or "—" }}</td>
              <td class="text-truncate" style="max-width:220px" title="{{ r.request_text }}">
                {{ r.request_text or "—" }}
              </td>
              <td class="text-center">
                {% if r.response_time_ms %}
                  {% set secs = r.response_time_ms / 1000 %}
                  <span class="{% if secs > 10 %}text-danger{% elif secs > 5 %}text-warning{% else %}text-success{% endif %}">
                    {{ "%.1f"|format(secs) }}s
                  </span>
                {% else %}—{% endif %}
              </td>
              <td class="text-center">
                {% if r.sentiment_label == 'positive' %}<span class="badge badge-pos">😊</span>
                {% elif r.sentiment_label == 'negative' %}<span class="badge badge-neg">😞</span>
                {% else %}<span class="badge badge-neu">😐</span>{% endif %}
              </td>
              <td class="small text-muted">{{ categorise_tools(r.tools_used) }}</td>
            </tr>
            {% else %}
            <tr><td colspan="7" class="text-center text-muted py-3">No requests in this period</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

</div><!-- /container -->

<script>
const BRAND = '#00796b'
const PALETTE = ['#00796b','#26a69a','#80cbc4','#004d40','#b2dfdb','#e0f2f1']

// Daily volume — line
const dailyData = {{ daily | tojson }}
new Chart(document.getElementById('lineChart'), {
  type: 'line',
  data: {
    labels: dailyData.map(d => d.day),
    datasets: [{
      label: 'Requests',
      data: dailyData.map(d => d.requests),
      borderColor: BRAND,
      backgroundColor: 'rgba(0,121,107,.1)',
      tension: .35,
      fill: true,
      pointBackgroundColor: BRAND,
    }]
  },
  options: { plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } }
})

// Agent bar
const agentData = {{ agents | tojson }}
new Chart(document.getElementById('agentBar'), {
  type: 'bar',
  data: {
    labels: agentData.map(a => a.agent_name),
    datasets: [
      { label: 'Requests', data: agentData.map(a => a.requests),
        backgroundColor: BRAND, borderRadius: 6 },
      { label: 'Conversations', data: agentData.map(a => a.sessions),
        backgroundColor: '#80cbc4', borderRadius: 6 },
    ]
  },
  options: { plugins: { legend: { position: 'bottom' } }, scales: { y: { beginAtZero: true } } }
})

// Sentiment doughnut
new Chart(document.getElementById('sentimentDoughnut'), {
  type: 'doughnut',
  data: {
    labels: ['Positive', 'Neutral', 'Negative'],
    datasets: [{
      data: [{{ kpi.pos_count or 0 }}, {{ kpi.neu_count or 0 }}, {{ kpi.neg_count or 0 }}],
      backgroundColor: ['#4caf50','#ffc107','#f44336'],
      borderWidth: 0,
    }]
  },
  options: { plugins: { legend: { position: 'bottom' } }, cutout: '65%' }
})

// Request type bar
const typeData = {{ type_counts | tojson }}
new Chart(document.getElementById('typeBar'), {
  type: 'bar',
  data: {
    labels: Object.keys(typeData),
    datasets: [{
      label: 'Requests',
      data: Object.values(typeData),
      backgroundColor: PALETTE,
      borderRadius: 6,
    }]
  },
  options: {
    indexAxis: 'y',
    plugins: { legend: { display: false } },
    scales: { x: { beginAtZero: true } }
  }
})

// Peak hours bar
const hourRaw = {{ hours | tojson }}
const hourLabels = Array.from({length:24}, (_,i) => i + ':00')
const hourVals   = hourLabels.map((_, i) => hourRaw[i] || 0)
new Chart(document.getElementById('hourBar'), {
  type: 'bar',
  data: {
    labels: hourLabels,
    datasets: [{ label: 'Requests', data: hourVals,
      backgroundColor: '#26a69a', borderRadius: 4 }]
  },
  options: { plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } }
})
</script>
</body>
</html>
"""

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == SUPERVISOR_PASSWORD:
            session["supervisor"] = True
            return redirect(url_for("dashboard"))
        error = "Incorrect password."
    return render_template_string(_LOGIN_HTML, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard():
    if not logged_in():
        return redirect(url_for("login"))
    week_offset = int(request.args.get("week", 0))
    start, end = _week_bounds(week_offset)
    stats = get_stats(start, end)
    return render_template_string(
        _DASHBOARD_HTML,
        start=start,
        end=end,
        week_offset=week_offset,
        categorise_tools=categorise_tools,
        **stats,
    )


def _build_csv_rows(start: str, end: str):
    stats = get_stats(start, end)
    rows = []
    rows.append(["LEADWAY HEALTH — WEEKLY REPORT", f"{start[:10]} to {end[:10]}"])
    rows.append([])
    rows.append(["SUMMARY"])
    rows.append(["Total Conversations", stats["kpi"].get("total_sessions", 0)])
    rows.append(["Total Requests",      stats["kpi"].get("total_requests", 0)])
    rows.append(["Avg Response Time (s)", stats["kpi"].get("avg_rt_s", 0)])
    rows.append(["Avg Sentiment Score",   stats["kpi"].get("avg_sentiment", 0)])
    rows.append(["Positive",  stats["kpi"].get("pos_count", 0)])
    rows.append(["Neutral",   stats["kpi"].get("neu_count", 0)])
    rows.append(["Negative",  stats["kpi"].get("neg_count", 0)])
    rows.append([])
    rows.append(["AGENT PERFORMANCE"])
    rows.append(["Agent", "Conversations", "Requests", "Avg Response Time (s)", "Avg Sentiment", "Escalations"])
    for a in stats["agents"]:
        rows.append([a["agent_name"], a["sessions"], a["requests"],
                     a["avg_rt_s"], a["avg_sentiment"], a["escalations"]])
    rows.append([])
    rows.append(["REQUEST LOG"])
    rows.append(["Time (UTC)", "Agent", "Member", "Request", "Response Time (ms)",
                 "Sentiment", "Service Type"])
    for r in stats["log"]:
        rows.append([
            r["requested_at"][:16].replace("T", " "),
            r["agent_name"],
            r["member_name"] or r["member_phone"] or "",
            r["request_text"] or "",
            r["response_time_ms"] or "",
            r["sentiment_label"] or "",
            categorise_tools(r["tools_used"]),
        ])
    return rows, start, end


@app.route("/download/csv")
def download_csv():
    if not logged_in():
        return redirect(url_for("login"))
    week_offset = int(request.args.get("week", 0))
    start, end = _week_bounds(week_offset)
    rows, s, e = _build_csv_rows(start, end)
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    filename = f"leadway_report_{s[:10]}_{e[:10]}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/download/excel")
def download_excel():
    if not logged_in():
        return redirect(url_for("login"))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    week_offset = int(request.args.get("week", 0))
    start, end = _week_bounds(week_offset)
    rows, s, e = _build_csv_rows(start, end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    HEADER_FILL = PatternFill("solid", fgColor="00796B")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SECTION_FONT = Font(bold=True, color="004D40")

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Style section headers
            if len(row) == 1 and isinstance(val, str) and val.isupper():
                cell.font = SECTION_FONT
            # Style column headers (rows after section headers)
            if r_idx > 1 and isinstance(rows[r_idx - 2], list) and \
               len(rows[r_idx - 2]) == 1 and isinstance(rows[r_idx - 2][0], str) and \
               rows[r_idx - 2][0].isupper() and isinstance(val, str):
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"leadway_report_{s[:10]}_{e[:10]}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.getenv("DASHBOARD_PORT", 5050))
    print(f"Supervisor dashboard running at http://localhost:{port}/")
    print(f"Default password: {SUPERVISOR_PASSWORD}")
    app.run(host="0.0.0.0", port=port, debug=False)
